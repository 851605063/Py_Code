#!python3
# coding: utf-8

'''
利用Python操作Excel的模块XlsxWriter，可以操作多个工作表的文字、数字、公式、图表等。
XlsxWriter模块具有以下功能：
    100%兼容的Excel XLSX文件，支持Excel 2003、Excel 2007等版本；
    支持所有Excel单元格数据格式；
    单元格合并、批注、自动筛选、丰富多格式字符串等；
    支持工作表PNG、JPEG图像，自定义图表；
    内存优化模式支持写入大文件。
'''
from openpyxl.styles.builtins import percent

''' 此例实现插入文字（中英字符）、数字（求和计算）、图片、单元格格式等 '''
import xlsxwriter
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, colors, Alignment
# from openpyxl.styles.colors import RED
import warnings
import re

warnings.filterwarnings('ignore')
def funcation1():
    workbook = xlsxwriter.Workbook('test.xlsx')  # 创建一个Excel文件

    worksheet = workbook.add_worksheet()  # 创建一个工作表对象

    worksheet.set_column('A:A', 20)  # 设定第一列（A）宽度为20像素
    bold = workbook.add_format({'bold': True})  # 定义一个加粗的格式对象

    worksheet.write('A1', 'Hello')  # A1单元格写入'Hello'
    worksheet.write('A2', 'World', bold)  # A2单元格写入'World'并引用加粗格式对象bold
    worksheet.write('B2', u'中文测试', bold)  # B2单元格写入中文并引用加粗格式对象bold
    worksheet.write(2, 0, 32)  # 用行列表示法写入数字'32'与'35.5'
    worksheet.write(3, 0, 35.5)  # 行列表示法的单元格下标以0作为起始值，'3,0'等价于'A3'
    worksheet.write(4, 0, '=SUM(A3:A4)')  # 求A3:A4的和，并将结果写入'4，0'，即'A5'
    # worksheet.insert_image('B5', 'img/python-logo.png')  # 在B5单元格插入图片
    workbook.close()  # 关闭Excel文件

def funcation2():
    excel=load_workbook(r"E:\wechat file\WeChat Files\wxid_welemk1qwdgj22\FileStorage\File\2021-01\UAT3预准出数据.xlsx")
    sheet=excel.get_sheet_by_name('Sheet1')
    wl = excel.create_sheet('网联')
    for xtc in range(1,sheet.max_column):
        title = sheet.cell(row=1, column=xtc).value
        if title=='系统':
            break
    print('xtc',xtc)

    for xtr in range(1,sheet.max_row):
        xt=sheet.cell(row=xtr,column=xtc).value
        if re.search('网联',xt):
            break
    print('xtr',xtr)
    #项目头写入
    for itemt in range(1, sheet.max_column):
        item1 = sheet.cell(row=1, column=itemt).value
        print(item1,' ',end='')
        wl.cell(row=1,column=itemt).value=item1
    #项目值写入
    bold_itatic_24_font = Font(name='等线', italic=False, color=colors.RED, bold=True)
    for itemz in range(1, sheet.max_column):
        item2 = sheet.cell(row=xtr, column=itemz).value
        print(item2,' ',end='')
        wl.cell(row=2, column=itemz).value=item2
        wl.cell(row=2, column=itemz).alignment = Alignment(horizontal='center', vertical='center')
        wl.cell(row=2, column=itemz).font=bold_itatic_24_font
    # wl.cell(row=2, column=10).style = percent
    #保存文件
    excel.save(r'C:\Users\85160\OneDrive\桌面\test.xlsx')





if __name__ == '__main__':
    funcation2()

